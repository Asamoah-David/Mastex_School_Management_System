"""
Export utilities for Mastex SchoolOS.
Provides streaming CSV, Excel, and ZIP export functionality.
"""
import csv
import io
import zipfile
from datetime import datetime
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone


class _Echo:
    """Pseudo-buffer that csv.writer can write to for streaming."""
    def write(self, value):
        return value


def _resolve_value(obj, field):
    """Extract a display value from an object for a given field spec."""
    field_path = field[1] if isinstance(field, tuple) else field
    if isinstance(obj, dict):
        value = obj.get(field_path, "")
    else:
        value = obj
        for part in field_path.split("__"):
            if hasattr(value, part):
                value = getattr(value, part)
            else:
                value = ""
                break
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _make_headers(fields):
    return [f[0] if isinstance(f, tuple) else f.replace("_", " ").title() for f in fields]


def export_to_csv(queryset, fields, filename=None):
    """Stream queryset as CSV (constant memory regardless of row count)."""
    if filename is None:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    echo = _Echo()
    writer = csv.writer(echo)

    def rows():
        yield writer.writerow(_make_headers(fields))
        for obj in queryset.iterator() if hasattr(queryset, "iterator") else queryset:
            yield writer.writerow([_resolve_value(obj, f) for f in fields])

    response = StreamingHttpResponse(rows(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_to_excel(queryset, fields, filename=None, sheet_name='Data'):
    """
    Export queryset to Excel file (.xlsx).
    
    Requires openpyxl to be installed.
    
    Args:
        queryset: Django queryset to export (can also be a list of dicts)
        fields: List of field names or (header, field_path) tuples
        filename: Optional filename
        sheet_name: Name of the Excel sheet
    
    Returns:
        HttpResponse with Excel content
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_to_csv(queryset, fields, filename)
    
    if filename is None:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write header
    headers = []
    for i, field in enumerate(fields, 1):
        if isinstance(field, tuple):
            headers.append(field[0])
        else:
            headers.append(field.replace('_', ' ').title())
        ws.cell(row=1, column=i, value=headers[-1])
    
    iterable = queryset.iterator() if hasattr(queryset, "iterator") else queryset
    for row_num, obj in enumerate(iterable, 2):
        for col_num, field in enumerate(fields, 1):
            ws.cell(row=row_num, column=col_num, value=_resolve_value(obj, field))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


def export_to_zip(exports, filename=None):
    """
    Export multiple querysets to a ZIP file.
    
    Args:
        exports: List of dicts with keys: queryset, fields, name, format ('csv' or 'excel')
        filename: Optional zip filename
    
    Returns:
        HttpResponse with ZIP content
    """
    if filename is None:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    with zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED) as zf:
        for export in exports:
            queryset = export['queryset']
            fields = export['fields']
            name = export.get('name', 'export')
            fmt = export.get('format', 'csv')
            
            if fmt == 'excel':
                content = export_to_excel(queryset, fields, f"{name}.xlsx")
            else:
                content = export_to_csv(queryset, fields, f"{name}.csv")
            
            # Read content and add to zip
            zf.writestr(f"{name}.{fmt}", content.content)
    
    return response


# ==================== COMMON FIELD DEFINITIONS ====================

STUDENT_FIELDS = [
    ('Admission No.', 'admission_number'),
    ('First Name', 'user__first_name'),
    ('Last Name', 'user__last_name'),
    ('Email', 'user__email'),
    ('Class', 'class_name'),
    ('Gender', 'gender'),
    ('Date of Birth', 'date_of_birth'),
    ('Status', 'status'),
]

STAFF_FIELDS = [
    ('Employee ID', 'employee_id'),
    ('First Name', 'user__first_name'),
    ('Last Name', 'user__last_name'),
    ('Email', 'user__email'),
    ('Role', 'role'),
    ('Phone', 'phone'),
    ('Status', 'status'),
]

FEE_FIELDS = [
    ('Admission No.', 'student__admission_number'),
    ('Student', 'student__user__first_name'),
    ('Class', 'student__class_name'),
    ('Total Amount', 'amount'),
    ('Paid', 'amount_paid'),
    ('Remaining', 'remaining_balance'),
    ('Status', 'payment_status'),
    ('Term', 'term'),
]

ATTENDANCE_FIELDS = [
    ('Date', 'date'),
    ('Student', 'student__user__first_name'),
    ('Admission No.', 'student__admission_number'),
    ('Class', 'student__class_name'),
    ('Status', 'status'),
    ('Remarks', 'remarks'),
]

EXPENSE_FIELDS = [
    ('Date', 'date'),
    ('Category', 'category__name'),
    ('Description', 'description'),
    ('Amount', 'amount'),
    ('Status', 'status'),
    ('Recorded By', 'recorded_by__first_name'),
]

INVENTORY_FIELDS = [
    ('Name', 'name'),
    ('Category', 'category__name'),
    ('Quantity', 'quantity'),
    ('Unit Price', 'unit_price'),
    ('Status', 'status'),
]