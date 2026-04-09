"""CSV/Excel result import wizard — upload, map columns, preview, confirm."""
import csv
import io
import json
import logging

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect

from accounts.decorators import role_required
from core.utils import get_school as _get_school
from students.models import Student
from .models import Subject, ExamType, Term, Result

logger = logging.getLogger(__name__)


def _parse_upload(file_obj):
    """Return (headers, rows) from a CSV or Excel upload."""
    name = file_obj.name.lower()
    if name.endswith((".xlsx", ".xls")):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return [], []
            headers = [str(h or "").strip() for h in rows[0]]
            data = [list(r) for r in rows[1:] if any(c is not None for c in r)]
            return headers, data
        except Exception as exc:
            logger.warning("Failed to parse Excel: %s", exc)
            return [], []
    else:
        try:
            text = file_obj.read().decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return [], []
            headers = [h.strip() for h in rows[0]]
            data = [r for r in rows[1:] if any(c.strip() for c in r)]
            return headers, data
        except Exception as exc:
            logger.warning("Failed to parse CSV: %s", exc)
            return [], []


COMMON_NAME_HEADERS = {"name", "student name", "student", "full name", "fullname"}
COMMON_ADMNO_HEADERS = {"admission number", "admission no", "admno", "adm no", "adm_no", "admission_number", "reg no", "reg_no"}
COMMON_SCORE_HEADERS = {"score", "mark", "marks", "total", "grade", "result", "total score"}
COMMON_REMARKS_HEADERS = {"remark", "remarks", "comment", "comments"}


def _auto_map(headers):
    """Return best-guess column mapping dict."""
    mapping = {"name_col": "", "admno_col": "", "score_col": "", "remarks_col": ""}
    lower = {i: h.lower().strip() for i, h in enumerate(headers)}
    for idx, h in lower.items():
        if h in COMMON_ADMNO_HEADERS:
            mapping["admno_col"] = str(idx)
        elif h in COMMON_NAME_HEADERS:
            mapping["name_col"] = str(idx)
        elif h in COMMON_SCORE_HEADERS:
            mapping["score_col"] = str(idx)
        elif h in COMMON_REMARKS_HEADERS:
            mapping["remarks_col"] = str(idx)
    return mapping


@role_required("school_admin", "teacher", "deputy_head", "hod")
def result_import_upload(request):
    """Step 1: Upload file + select class/subject/exam/term."""
    school = _get_school(request)
    if not school:
        return redirect("home")

    classes = list(Student.objects.filter(school=school).values_list("class_name", flat=True).distinct())
    classes = sorted(c for c in classes if c)
    subjects = Subject.objects.filter(school=school).order_by("name")
    exam_types = ExamType.objects.filter(school=school).order_by("name")
    terms = Term.objects.filter(school=school).order_by("-is_current", "-id")
    current_term = terms.filter(is_current=True).first()

    if request.method == "POST":
        file = request.FILES.get("file")
        class_name = request.POST.get("class_name", "").strip()
        subject_id = request.POST.get("subject")
        exam_type_id = request.POST.get("exam_type")
        term_id = request.POST.get("term")

        if not file or not class_name or not subject_id or not exam_type_id or not term_id:
            messages.error(request, "All fields are required.")
            return redirect("academics:result_import_upload")

        headers, rows = _parse_upload(file)
        if not headers or not rows:
            messages.error(request, "Could not read data from the uploaded file.")
            return redirect("academics:result_import_upload")

        request.session["import_headers"] = headers
        request.session["import_rows"] = [[str(c) if c is not None else "" for c in r] for r in rows]
        request.session["import_meta"] = {
            "class_name": class_name,
            "subject_id": subject_id,
            "exam_type_id": exam_type_id,
            "term_id": term_id,
        }
        return redirect("academics:result_import_map")

    return render(request, "academics/result_import_upload.html", {
        "school": school,
        "classes": classes,
        "subjects": subjects,
        "exam_types": exam_types,
        "terms": terms,
        "current_term_id": str(current_term.id) if current_term else None,
    })


@role_required("school_admin", "teacher", "deputy_head", "hod")
def result_import_map(request):
    """Step 2: Map file columns to student/score fields."""
    school = _get_school(request)
    if not school:
        return redirect("home")

    headers = request.session.get("import_headers")
    rows = request.session.get("import_rows")
    meta = request.session.get("import_meta")
    if not headers or not rows or not meta:
        messages.error(request, "No uploaded file found. Please start over.")
        return redirect("academics:result_import_upload")

    auto = _auto_map(headers)

    if request.method == "POST":
        mapping = {
            "name_col": request.POST.get("name_col", ""),
            "admno_col": request.POST.get("admno_col", ""),
            "score_col": request.POST.get("score_col", ""),
            "remarks_col": request.POST.get("remarks_col", ""),
        }
        if not mapping["score_col"]:
            messages.error(request, "Score column is required.")
            return render(request, "academics/result_import_map.html", {
                "school": school, "headers": headers, "preview_rows": rows[:5], "auto": auto,
            })
        request.session["import_mapping"] = mapping
        return redirect("academics:result_import_preview")

    return render(request, "academics/result_import_map.html", {
        "school": school,
        "headers": headers,
        "preview_rows": rows[:5],
        "auto": auto,
    })


@role_required("school_admin", "teacher", "deputy_head", "hod")
def result_import_preview(request):
    """Step 3: Preview matched rows + validate."""
    school = _get_school(request)
    if not school:
        return redirect("home")

    headers = request.session.get("import_headers")
    rows = request.session.get("import_rows")
    meta = request.session.get("import_meta")
    mapping = request.session.get("import_mapping")
    if not all([headers, rows, meta, mapping]):
        messages.error(request, "Session expired. Please start over.")
        return redirect("academics:result_import_upload")

    score_idx = int(mapping["score_col"]) if mapping["score_col"] else None
    admno_idx = int(mapping["admno_col"]) if mapping["admno_col"] else None
    name_idx = int(mapping["name_col"]) if mapping["name_col"] else None
    remarks_idx = int(mapping["remarks_col"]) if mapping["remarks_col"] else None

    students_in_class = {
        s.admission_number.strip().lower(): s
        for s in Student.objects.filter(school=school, class_name=meta["class_name"]).select_related("user")
        if s.admission_number
    }
    students_by_name = {}
    for s in Student.objects.filter(school=school, class_name=meta["class_name"]).select_related("user"):
        full = s.user.get_full_name().strip().lower() if s.user else ""
        if full:
            students_by_name[full] = s

    preview = []
    for row in rows:
        entry = {"raw": row, "student": None, "score": None, "remarks": "", "error": ""}
        admno = row[admno_idx].strip() if admno_idx is not None and admno_idx < len(row) else ""
        name = row[name_idx].strip() if name_idx is not None and name_idx < len(row) else ""
        score_raw = row[score_idx].strip() if score_idx is not None and score_idx < len(row) else ""
        remarks = row[remarks_idx].strip() if remarks_idx is not None and remarks_idx < len(row) else ""

        matched = students_in_class.get(admno.lower()) if admno else None
        if not matched and name:
            matched = students_by_name.get(name.lower())
        entry["student"] = matched

        if not matched:
            entry["error"] = "Student not found"
        try:
            s = float(score_raw)
            if 0 <= s <= 100:
                entry["score"] = s
            else:
                entry["error"] = "Score out of range (0-100)"
        except (ValueError, TypeError):
            entry["error"] = "Invalid score"
        entry["remarks"] = remarks
        preview.append(entry)

    valid_count = sum(1 for p in preview if p["student"] and p["score"] is not None and not p["error"])
    error_count = sum(1 for p in preview if p["error"])

    if request.method == "POST":
        return redirect("academics:result_import_confirm")

    return render(request, "academics/result_import_preview.html", {
        "school": school,
        "preview": preview,
        "valid_count": valid_count,
        "error_count": error_count,
        "meta": meta,
    })


@role_required("school_admin", "teacher", "deputy_head", "hod")
def result_import_confirm(request):
    """Step 4: Execute the import."""
    school = _get_school(request)
    if not school:
        return redirect("home")

    headers = request.session.get("import_headers")
    rows = request.session.get("import_rows")
    meta = request.session.get("import_meta")
    mapping = request.session.get("import_mapping")
    if not all([headers, rows, meta, mapping]):
        messages.error(request, "Session expired. Please start over.")
        return redirect("academics:result_import_upload")

    try:
        subject = Subject.objects.get(id=meta["subject_id"], school=school)
        exam_type = ExamType.objects.get(id=meta["exam_type_id"], school=school)
        term = Term.objects.get(id=meta["term_id"], school=school)
    except (Subject.DoesNotExist, ExamType.DoesNotExist, Term.DoesNotExist):
        messages.error(request, "Invalid subject/exam/term selection.")
        return redirect("academics:result_import_upload")

    score_idx = int(mapping["score_col"]) if mapping["score_col"] else None
    admno_idx = int(mapping["admno_col"]) if mapping["admno_col"] else None
    name_idx = int(mapping["name_col"]) if mapping["name_col"] else None
    remarks_idx = int(mapping["remarks_col"]) if mapping["remarks_col"] else None

    students_in_class = {
        s.admission_number.strip().lower(): s
        for s in Student.objects.filter(school=school, class_name=meta["class_name"]).select_related("user")
        if s.admission_number
    }
    students_by_name = {}
    for s in Student.objects.filter(school=school, class_name=meta["class_name"]).select_related("user"):
        full = s.user.get_full_name().strip().lower() if s.user else ""
        if full:
            students_by_name[full] = s

    saved = 0
    for row in rows:
        admno = row[admno_idx].strip() if admno_idx is not None and admno_idx < len(row) else ""
        name = row[name_idx].strip() if name_idx is not None and name_idx < len(row) else ""
        score_raw = row[score_idx].strip() if score_idx is not None and score_idx < len(row) else ""
        remarks = row[remarks_idx].strip() if remarks_idx is not None and remarks_idx < len(row) else ""

        matched = students_in_class.get(admno.lower()) if admno else None
        if not matched and name:
            matched = students_by_name.get(name.lower())
        if not matched:
            continue
        try:
            score = float(score_raw)
            if not (0 <= score <= 100):
                continue
        except (ValueError, TypeError):
            continue

        Result.objects.update_or_create(
            student=matched, subject=subject, exam_type=exam_type, term=term,
            defaults={"score": score},
        )
        saved += 1

    for key in ("import_headers", "import_rows", "import_meta", "import_mapping"):
        request.session.pop(key, None)

    messages.success(request, f"Successfully imported {saved} result(s).")
    return redirect("academics:result_list")
